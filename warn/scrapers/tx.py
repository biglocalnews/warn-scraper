import logging
import re
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .. import utils
from ..cache import Cache

__authors__ = ["Dilcia19", "ydoc5212"]
__tags__ = ["html", "excel", "historical"]
__source__ = {
    "name": "Texas Workforce Commission",
    "url": "https://www.twc.texas.gov/businesses/worker-adjustment-and-retraining-notification-warn-notices#warnNotices",
}

logger = logging.getLogger(__name__)


def scrape(
    data_dir: Path = utils.WARN_DATA_DIR,
    cache_dir: Path = utils.WARN_CACHE_DIR,
) -> Path:
    """
    Scrape data from Texas.

    Keyword arguments:
    data_dir -- the Path were the result will be saved (default WARN_DATA_DIR)
    cache_dir -- the Path where results can be cached (default WARN_CACHE_DIR)

    Returns: the Path where the file is written
    """
    # Set up the cache
    cache = Cache(cache_dir)

    # Get the root URL
    url = "https://www.twc.texas.gov/businesses/worker-adjustment-and-retraining-notification-warn-notices#warnNotices"
    page = utils.get_url(url)
    html = page.text

    # Cache it
    cache.write("tx/source.html", html)

    # Get all the Excel links
    soup = BeautifulSoup(page.text, "html5lib")
    link_list = soup.find_all("a", href=re.compile("^/files/news/warn-act-listings-"))

    # Clean up the links and filter 'em down
    href_list = []
    for link in link_list:
        # Extract year
        href = link.get("href")
        # Only keep links after 2019 since our historical document covers 2018 and before
        year = _get_year(href)
        if year >= 2019:
            href_list.append(href)

    # Loop through the links we want to download
    row_list = []
    for ihref, href in enumerate(href_list):

        # get each url from the HTML links we found
        data_url = f"https://www.twc.texas.gov{href}"

        # download the excel file
        year = _get_year(href)
        ext = _get_ext(href)
        excel_path = cache.download(f"tx/{year}{ext}", data_url)

        # Open it up
        workbook = load_workbook(filename=excel_path)

        # Get the first sheet
        worksheet = workbook.worksheets[0]

        # Convert the sheet to a list of lists
        for irow, row in enumerate(worksheet.rows):

            # Skip headers after the first workbook
            if ihref > 0 and irow == 0:
                continue
            cell_list = [cell.value for cell in row]

            # Skip empty rows
            if cell_list[0] is None:
                continue

            # Add what's left to the pile
            row_list.append(cell_list)

    # Get historical URL
    historical_url = (
        "https://storage.googleapis.com/bln-data-public/warn-layoffs/tx_historical.xlsx"
    )
    excel_path = cache.download("tx/historical.xlsx", historical_url)

    # Open it up
    workbook = load_workbook(filename=excel_path)

    # Get the first sheet
    worksheet = workbook.worksheets[0]

    # Convert the sheet to a list of lists
    for i, row in enumerate(worksheet.rows):
        # Skip header
        if i == 0:
            continue

        # Trim down to only the columns in the scrape, so they match
        select_columns = [
            row[8],  # NOTICE_DATE
            row[0],  # JOB_SITE_NAME
            row[2],  # COUNTY_NAME
            row[5],  # WDA_NAME
            row[6],  # TOTAL_LAYOFF_NUMBER
            row[7],  # LayOff_Date
            row[11],  # WFDD_RECEIVED_DATE
            row[1],  # CITY_NAME
        ]

        # Tack 'em on
        cell_list = [c.value for c in select_columns]
        row_list.append(cell_list)

    # Set the export path
    data_path = data_dir / "tx.csv"

    # Write out the file
    utils.write_rows_to_csv(data_path, row_list)

    # Return the path to the file
    return data_path


def _get_year(url: str) -> int:
    """Plucks the year from the provided URL."""
    filename_regex = re.match(r".*-(\d{4})(.*)$", url, re.I)
    assert filename_regex is not None
    year_str = filename_regex.group(1)[-4:]
    return int(year_str)


def _get_ext(url: str) -> str:
    """Plucks the file extension from the provided URL."""
    filename_regex = re.match(r".*-(\d{4})(.*)$", url, re.I)
    assert filename_regex is not None
    return filename_regex.group(2)


if __name__ == "__main__":
    scrape()
