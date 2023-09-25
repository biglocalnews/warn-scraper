import logging
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .. import utils
from ..cache import Cache

__authors__ = ["zstumgoren", "Dilcia19", "ydoc5212", "stucka"]
__tags__ = ["historical", "excel"]
__source__ = {
    "name": "Oregon Higher Education Coordinating Commission",
    "url": "https://ccwd.hecc.oregon.gov/Layoff/WARN",
}

logger = logging.getLogger(__name__)


def scrape(
    data_dir: Path = utils.WARN_DATA_DIR,
    cache_dir: Path = utils.WARN_CACHE_DIR,
) -> Path:
    """
    Scrape data from Oregon.

    Keyword arguments:
    data_dir -- the Path were the result will be saved (default WARN_DATA_DIR)
    cache_dir -- the Path where results can be cached (default WARN_CACHE_DIR)

    Returns: the Path where the file is written
    """
    # Initialize the cache
    cache = Cache(cache_dir)

    starturl = "https://ccwd.hecc.oregon.gov/Layoff/WARN/Download"
    baseurl = "https://ccwd.hecc.oregon.gov"

    r = requests.get(starturl)

    cookies = r.cookies

    soup = BeautifulSoup(r.content, features="html5lib")

    # Looking for something like <input name="__RequestVerificationToken" type="hidden" value="GYlfHSHzATg5x9TZgIe...
    tokenname = "__RequestVerificationToken"
    tokenvalue = soup.find("input", {"name": tokenname})["value"]

    payload = {
        tokenname: tokenvalue,
        "WARNFormat": "xlsx",
        "WARNSort": "LOT4",
    }

    requestheaders = {
        "Host": "ccwd.hecc.oregon.gov",
        "User-Agent": "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:109.0) Gecko/20100101 Firefox/116.0",
        "Origin": "https://ccwd.hecc.oregon.gov",
        "Referer": "https://ccwd.hecc.oregon.gov/Layoff/WARN/Download",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-User": "?1",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.5",
        "Accept-Encoding": "gzip, deflate, br",
        "Content-Type": "application/x-www-form-urlencoded",
        "Connection": "keep-alive",
    }

    r = requests.post(starturl, cookies=cookies, data=payload, headers=requestheaders)

    dlsoup = BeautifulSoup(r.content, features="html5lib")
    excelurl = (
        baseurl + dlsoup.find("a", {"target": "_blank", "class": "btn-primary"})["href"]
    )
    logger.debug(f"Found latest data's URL at {excelurl}")
    if not excelurl:
        logger.error("No URL could be found for the newest spreadsheet.")
    latest_excel_path = "or/latest.xlsx"
    logger.debug(f"Trying to save to, we hope, {cache_dir/latest_excel_path}")
    cache.download(latest_excel_path, excelurl)

    workbook = load_workbook(filename=cache_dir/latest_excel_path)
    worksheet = workbook.worksheets[0]

    masterlist: list = []
    headers: list = []

    sheetrows = list(worksheet.rows)
    for cell in sheetrows[2]:
        headers.append(cell.value)
    for row in sheetrows[3:]:
        line = {}
        for i, item in enumerate(headers):
            line[item] = list(row)[i].value
        if (
            len(line[headers[0]]) + len(line[headers[1]])
        ) != 0:  # Filter out blank rows
            masterlist.append(line)
    historicalurl = (
        "https://storage.googleapis.com/bln-data-public/warn-layoffs/or_historical.xlsx"
    )
    historical_excel_path = str(cache_dir) + "/or/historical.xlsx"

    utils.fetch_if_not_cached(historical_excel_path, historicalurl)
    workbook = load_workbook(filename=historical_excel_path)

    # Get the first sheet
    worksheet = workbook.worksheets[0]
    sheetrows = list(worksheet.rows)

    historical_headers = []
    for cell in sheetrows[2]:
        historical_headers.append(cell.value)

    if historical_headers != headers:
        logger.error("Newest headers no longer match historical headers")
    else:
        logger.debug("OK! Newest headers match historical headers.")

    duplicated_rows = 0
    for row in sheetrows[3:]:
        line = {}
        for i, item in enumerate(headers):
            line[item] = list(row)[i].value
        if (
            len(line[headers[0]]) + len(line[headers[1]])
        ) != 0:  # Filter out blank rows
            if line in masterlist:
                duplicated_rows += 1
            else:
                masterlist.append(line)

    logger.debug(f"{duplicated_rows:,} duplicated rows not added.")

    data_path = data_dir / "or.csv"
    utils.write_dict_rows_to_csv(
        data_path, headers, masterlist, mode="w", extrasaction="raise"
    )

    # Return the path to the file
    return data_path


if __name__ == "__main__":
    scrape()
