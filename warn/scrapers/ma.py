import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urljoin

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from .. import utils
from ..cache import Cache

__authors__ = ["riordan"]
__tags__ = ["excel", "csv", "historical"]
__source__ = {
    "name": "Massachusetts Executive Office of Labor and Workforce Development",
    "url": "https://www.mass.gov/info-details/worker-adjustment-and-retraining-notification-act-warn-layoff-and-closure-updates",
}

logger = logging.getLogger(__name__)

BASE_URL = "https://www.mass.gov"

# The unified header for the combined export. It mirrors the columns the state
# currently publishes in its flat fiscal-year workbooks and its weekly CSV feed.
CANONICAL_HEADER = [
    "RECEIVED",
    "EMPLOYER",
    "CITY/TOWN",
    "REGION",
    "DATE(S) OF LAYOFFS",
    "# EMPLOYEES IMPACTED",
]

# The older fiscal-year workbooks (FY2022, FY2023) split notices into one
# worksheet per region and lay the columns out in this fixed order. The region
# for those rows comes from the worksheet name rather than a column.
REGION_SHEET_ORDER = [
    "RECEIVED",
    "EMPLOYER",
    "CITY/TOWN",
    "DATE(S) OF LAYOFFS",
    "# EMPLOYEES IMPACTED",
]


def scrape(
    data_dir: Path = utils.WARN_DATA_DIR,
    cache_dir: Path = utils.WARN_CACHE_DIR,
) -> Path:
    """
    Scrape data from Massachusetts.

    Keyword arguments:
    data_dir -- the Path were the result will be saved (default WARN_DATA_DIR)
    cache_dir -- the Path where results can be cached (default WARN_CACHE_DIR)

    Returns: the Path where the file is written
    """
    # Fire up the cache
    cache = Cache(cache_dir)
    state_code = "ma"

    # Download the index page. Mass.gov sits behind Akamai bot protection that
    # rejects ordinary requests, so we route everything through Zyte, as the
    # Louisiana and Texas scrapers do.
    _, html = utils.get_with_zyte(__source__["url"])
    cache.write(f"{state_code}/index.html", html)

    # Find the links to the fiscal-year workbooks and the weekly CSV feed.
    soup = BeautifulSoup(html, "html.parser")
    excel_urls, csv_urls = _find_source_links(soup)
    logger.debug(
        f"Found {len(excel_urls)} workbook link(s) and {len(csv_urls)} CSV link(s)"
    )

    master_list: list = []

    # The fiscal-year workbooks hold the historical record.
    for url in excel_urls:
        name = url.split("/doc/")[1].split("/")[0]
        raw, _ = utils.get_with_zyte(url)
        excel_path = cache.write_binary(f"{state_code}/{name}.xlsx", raw)
        master_list.extend(_parse_workbook(excel_path))

    # The weekly CSV feed carries the current fiscal year's running notices,
    # which are not yet folded into a workbook.
    for url in csv_urls:
        name = os.path.basename(unquote(url))
        raw, text = utils.get_with_zyte(url)
        cache.write_binary(f"{state_code}/{name}", raw)
        rows = list(csv.reader(text.splitlines()))
        master_list.extend(_parse_flat_rows(rows))

    # Drop exact duplicate rows while preserving order. Duplicates can appear at
    # fiscal-year boundaries or across re-runs; distinct notices are untouched.
    deduped = _dedupe(master_list)
    logger.debug(
        f"Parsed {len(master_list):,} rows, {len(deduped):,} after removing duplicates"
    )

    # Write it out
    data_path = data_dir / f"{state_code}.csv"
    utils.write_dict_rows_to_csv(
        data_path, CANONICAL_HEADER, deduped, extrasaction="raise"
    )

    return data_path


def _find_source_links(soup: BeautifulSoup) -> tuple:
    """Pull the workbook and CSV download links off the index page.

    Args:
        soup (BeautifulSoup): the parsed index page

    Returns: a tuple of (sorted workbook urls, sorted csv urls)
    """
    excel_urls = set()
    csv_urls = set()
    for link in soup.find_all("a", href=True):
        href = link["href"]
        lower = href.lower()
        # Fiscal-year workbooks live at /doc/fyNN-warn-report.../download
        if "/doc/" in lower and "warn-report" in lower:
            excel_urls.add(urljoin(BASE_URL, href))
        # The weekly feed is a WARN Report CSV under /files/
        elif lower.endswith(".csv") and "warn" in lower:
            csv_urls.add(urljoin(BASE_URL, href))
    return sorted(excel_urls), sorted(csv_urls)


def _parse_workbook(excel_path: Path) -> list:
    """Parse a fiscal-year workbook into a list of canonical row dicts.

    Args:
        excel_path (Path): the path to a cached .xlsx file

    Returns: a list of dicts keyed by CANONICAL_HEADER
    """
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    rows: list = []
    for sheet in workbook.worksheets:
        sheet_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        if _has_region_column(sheet_rows):
            rows.extend(_parse_flat_rows(sheet_rows))
        else:
            rows.extend(_parse_region_sheet(sheet_rows, sheet.title.strip()))
    workbook.close()
    return rows


def _has_region_column(rows: list) -> bool:
    """Return True if the sheet is a flat table with its own REGION column."""
    header_index = _find_header_row(rows)
    if header_index is None:
        return False
    return any(_norm(cell) == "REGION" for cell in rows[header_index])


def _find_header_row(rows: list):
    """Find the index of the header row, or None if there isn't one."""
    for i, row in enumerate(rows):
        labels = {_norm(cell) for cell in row}
        if "EMPLOYER" in labels or "COMPANY NAME" in labels:
            return i
    return None


def _parse_flat_rows(rows: list) -> list:
    """Parse a flat table (FY2024+ workbooks and the weekly CSV) into row dicts.

    The header row names its own columns, including REGION, so we map each
    canonical field to the column where its header appears.
    """
    header_index = _find_header_row(rows)
    if header_index is None:
        return []

    # Map each canonical column name to its position in the header row.
    positions = {}
    for col, cell in enumerate(rows[header_index]):
        label = _norm(cell)
        if label in CANONICAL_HEADER and label not in positions:
            positions[label] = col

    parsed = []
    for row in rows[header_index + 1 :]:
        record = {
            field: _clean(row[pos]) if pos < len(row) else ""
            for field, pos in positions.items()
        }
        record = {field: record.get(field, "") for field in CANONICAL_HEADER}
        if _keep_row(record):
            parsed.append(record)
    return parsed


def _parse_region_sheet(rows: list, region: str) -> list:
    """Parse an older per-region worksheet (FY2022, FY2023) into row dicts.

    These sheets have no REGION column and lay their columns out positionally,
    so the region comes from the worksheet name.
    """
    header_index = _find_header_row(rows)
    if header_index is None:
        return []

    parsed = []
    for row in rows[header_index + 1 :]:
        record = {field: "" for field in CANONICAL_HEADER}
        for pos, field in enumerate(REGION_SHEET_ORDER):
            if pos < len(row):
                record[field] = _clean(row[pos])
        record["REGION"] = region
        if _keep_row(record):
            parsed.append(record)
    return parsed


def _keep_row(record: dict) -> bool:
    """Decide whether a parsed row is a real notice worth keeping."""
    employer = record.get("EMPLOYER", "")
    if not employer:
        return False
    # Skip stray header echoes and any total/count footer rows.
    if _norm(employer) in ("EMPLOYER", "COMPANY NAME"):
        return False
    if _norm(employer).startswith("TOTAL"):
        return False
    return True


def _dedupe(rows: list) -> list:
    """Drop exact duplicate rows while preserving their first-seen order."""
    seen = set()
    deduped = []
    for row in rows:
        key = tuple(row[field] for field in CANONICAL_HEADER)
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped


def _clean(value) -> str:
    """Normalize a cell value to a clean string for output.

    Dates arrive from openpyxl as datetime objects; render them the way the
    state writes them elsewhere (M/D/YYYY). Everything else is stringified and
    stripped.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    return str(value).strip()


def _norm(value) -> str:
    """Uppercase, whitespace-collapsed form of a cell, for label matching."""
    if value is None:
        return ""
    return " ".join(str(value).split()).upper()


if __name__ == "__main__":
    scrape()
